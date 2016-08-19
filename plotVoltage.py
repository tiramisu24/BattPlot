import openpyxl
import os
import plotGraph
from plotGraph import PlotGraph
from os import listdir
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pdb

class PlotVoltage(PlotGraph):
        
    def breakCycles(self, filename, sheetName, cycle,voltage, capacity, current, areaElectrode):
        dictCycles ={}

        temp = {}
        chargeCap = []
        chargeVol = []
        discharCap = []
        discharVol = []
        wb = load_workbook(filename)
        
        curSheet = wb[sheetName]            
        max_length = curSheet.max_row                  
        prevCycle = 1
        
        for row in range(3, max_length+1):
            try:
                curCycle = int(curSheet[cycle + str(row)].value)
                curRowCur = curSheet[current + str(row)].value
                curRowCap = float(curSheet[capacity + str(row)].value)/float(areaElectrode)
                curRowVol = curSheet[voltage + str(row)].value
                prevRowVol = curSheet[voltage + str(row-1)].value

                if row == max_length or curCycle>prevCycle:

                    temp['chargeCap']=chargeCap
                    temp['chargeVol']=chargeVol
                    temp['dischargeCap']=discharCap
                    temp['dischargeVol']=discharVol

                    dictCycles[prevCycle]=temp                

                    chargeCap = []
                    chargeVol = []
                    discharCap = []
                    discharVol = []
                    temp ={}
                    
                    prevCycle = curCycle       
                
                if curRowCur ==0 or curRowCap ==0:
                    continue
                
                if prevRowVol<=curRowVol and curRowCur >0:
                    chargeCap.append(curRowCap)
                    chargeVol.append(curRowVol)
                    
                elif prevRowVol>=curRowVol and curRowCur <0:
                    discharCap.append(curRowCap)
                    discharVol.append(curRowVol)
            except:
                break
        return dictCycles

    def plot_data(self, file_names, graphTitle, AreaElectrode, numCycles,
                  currentCol, cycleCol,sheetName ,voltage, capacity,
                  YAxisLimit, YAxisLower, XAxisLimit, XAxisLower, 
                  YaxisLabel,XaxisLabel):
        
        self.setParam(graphTitle, AreaElectrode, float(YAxisLimit), float(YAxisLower), 
                                            float(XAxisLimit), float(XAxisLower),YaxisLabel,XaxisLabel)     
        
        formatColors = ['g', 'r', 'k','c','m', 'y']
        n=0
        
        for filename in file_names:
            comp = filename[-4:]
            legendName = filename[-13:] 
            if (comp =='xlsx'):

                dictCycles = self.breakCycles(filename, sheetName ,cycleCol, voltage, capacity, currentCol, AreaElectrode)
            else:
                continue
            for cycle in numCycles:  
                cycle = int(cycle)                              
                tempDataList = dictCycles[cycle]
                line1, =plt.plot(tempDataList['chargeCap'],tempDataList['chargeVol'] , formatColors[n])
                line2, =plt.plot(tempDataList['dischargeCap'],tempDataList['dischargeVol'],formatColors[n])            
                line1.set_label(str(cycle) + " Charge " + legendName)
                line2.set_label(str(cycle) + " Discharge " + legendName)
                
                if n==5:
                    n=0
                else:               
                    n +=1
        
        plt.legend(loc ='center left', bbox_to_anchor =(1,0.5))            
        plt.show()
        
    