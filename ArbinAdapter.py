from openpyxl import load_workbook
import pdb

thisList = ['NCA_P_01.xlsx','NCA_P_02.xlsx','NCA_P_03.xlsx']

for fileName in thisList:
    wb = load_workbook(fileName)    
    cur_sheet = wb['Sheet1']
    max_length = cur_sheet.max_row
    
    rowNum =2
    while rowNum <max_length:
        dischargeData = cur_sheet['J' + str(rowNum)].value
        chargeData = cur_sheet['I' + str(rowNum)].value
        
        if dischargeData == 0:
            cur_sheet['J' + str(rowNum)] = chargeData    
        rowNum +=1
    
    wb.save('result' + fileName)
