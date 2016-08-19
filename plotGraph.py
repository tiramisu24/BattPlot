import os
from os import listdir
from openpyxl import load_workbook
import string
import matplotlib.pyplot as plt
import pdb


class PlotGraph():
    
    def setParam(self, graphTitle, AreaElectrode, YAxisLimit, YAxisLower, XAxisLimit, XAxisLower, YaxisLabel,XaxisLabel):
        plt.figure()
        plt.ylabel(YaxisLabel)
        plt.xlabel(XaxisLabel)
        plt.axis([XAxisLower,XAxisLimit,YAxisLower,YAxisLimit])
        plt.title(graphTitle) 
        self.AreaElectrode = AreaElectrode

    
    def get_names(self, address):
        try:
            if os.path.isdir(address):
                cur_list_dir = listdir(address)
            else:
                return "Invalid directory"
            
            self.name_list = cur_list_dir
            
            self.directory_path=address
            os.chdir(self.directory_path)
            
            return cur_list_dir    
        except OSError:
            return "Address is not a directory"   
    
    
    

    def get_sheetnames(self, listDataFiles):
        for item in listDataFiles:
            
            comp = item[-4:] 
            if (comp =='xlsx'):
                first_file = item
                break
        
        wb = load_workbook(first_file, data_only= True)
        
        sheet_name = wb.get_sheet_names()
        return sheet_name
        
    def get_headings(self, listDataFiles,sheetName):
        
        for item in listDataFiles:            
            comp = item[-4:] 
            if (comp =='xlsx'):
                first_file = item
                break
        
        wb = load_workbook(first_file, data_only= True)
        
        sheet_name = wb[sheetName]

        max_length = sheet_name.max_column
        alpha_list = list(string.ascii_uppercase) 
        count = 0
        first_row=[]
        first_row_dic={}
        
        while count < max_length:
            index_val = alpha_list[count] + str(2)
            cell_value = sheet_name[index_val].value           
            first_row.append(cell_value)
            first_row_dic[cell_value] = index_val
            count +=1         
         
        return first_row_dic
    
        

    def extract_data (self, tupleoftuple, AreaElectrode):
        listdata = []
        for item in tupleoftuple:
            cell_data = item[0].value
            try:
                cell_data = float(cell_data)/float(AreaElectrode)
            except ValueError:
                continue
            except TypeError:
                continue
            
            listdata.append(cell_data)                
        return listdata
        
    def closeGraph(self):
        plt.close()
