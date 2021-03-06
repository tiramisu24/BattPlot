import openpyxl
import os
from os import listdir
from openpyxl import load_workbook
import string
import matplotlib.pyplot as plt
from openpyxl import Workbook
import plotGraph
from plotGraph import PlotGraph
import pdb


class PlotCapacity(PlotGraph):

    def plot_data(self, file_names, sheet_num, column_cell, set_num,
                  graphTitle, AreaElectrode, 
                  YAxisLimit, YAxisLower, XAxisLimit, XAxisLower, 
                  YaxisLabel, XaxisLabel):
        
        count =0
        
        for file_name in file_names:
            legendName = file_name[-13:] 
            wb = load_workbook(file_name)
            cur_sheet = wb[sheet_num]
            max_length = cur_sheet.max_row
            end_cell = column_cell + str(max_length)
            start_cell = column_cell + str(1)
            temp_cell = cur_sheet[start_cell:end_cell]

            capacity = self.extract_data(temp_cell, AreaElectrode)  
            max_length = len(capacity)     
 
            cycle_number = range(1,max_length+1)
            if count >set_num or count == 0:
                if count != 0:
                    plt.legend(loc ='center left', bbox_to_anchor =(1,0.5))            

                self.setParam(graphTitle, AreaElectrode, float(YAxisLimit), float(YAxisLower), 
                                    float(XAxisLimit), float(XAxisLower),YaxisLabel,XaxisLabel)                
                count = 1                
                            
            line, = plt.plot(cycle_number,capacity)
            line.set_label(legendName)
            count +=1
            
        plt.legend(loc ='center left', bbox_to_anchor =(1,0.5))            
        plt.show()
        
 